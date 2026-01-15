from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.forms import modelformset_factory
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from apps.ticket.models import Ticket, TicketDetail
from apps.ticket.forms import TicketForm, TicketDetailForm


class TicketListView(ListView):
    """
    Vista para listar tickets con filtros y búsqueda.
    """
    model = Ticket
    template_name = 'ticket/ticket_list.html'
    context_object_name = 'tickets'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('company')
        seller = self.request.GET.get('seller')
        if seller:
            queryset = queryset.filter(seller__icontains=seller)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                client__icontains=search
            ) | queryset.filter(
                seller__icontains=search
            ) | queryset.filter(
                ci_ruc__icontains=search
            )
        return queryset.order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sellers'] = Ticket.objects.exclude(seller__isnull=True).exclude(seller='').values_list('seller', flat=True).distinct().order_by('seller')
        # Breadcrumbs
        context['breadcrumb_list'] = [
            {'label': 'Dashboard', 'url': reverse_lazy('core:dashboard')},
            {'label': 'Tickets'}
        ]
        return context


class TicketDetailView(DetailView):
    """
    Vista para ver detalles de un ticket específico.
    """
    model = Ticket
    template_name = 'ticket/ticket_detail.html'
    context_object_name = 'ticket'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['details'] = self.object.details.all()
        # Breadcrumbs
        context['breadcrumb_list'] = [
            {'label': 'Dashboard', 'url': reverse_lazy('core:dashboard')},
            {'label': 'Tickets', 'url': reverse_lazy('ticket:ticket_list')},
            {'label': f'Ticket #{self.object.document_number}'}
        ]
        return context


class TicketCreateView(CreateView):
    """
    Vista para crear un ticket con detalles usando formsets.
    """
    model = Ticket
    form_class = TicketForm
    template_name = 'ticket/ticket_form.html'
    success_url = reverse_lazy('ticket:ticket_list')

    def get(self, request, *args, **kwargs):
        from apps.company.models import Company
        if not Company.objects.exists():
            messages.warning(request, 'Debe crear al menos una compañía antes de crear tickets.')
            return redirect('company:company_list')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        TicketDetailFormSet = modelformset_factory(
            TicketDetail,
            form=TicketDetailForm,
            extra=1,  # Una fila extra para agregar
            can_delete=True,  # Permitir eliminar
        )
        if self.request.POST:
            context['detail_formset'] = TicketDetailFormSet(
                self.request.POST,
                queryset=TicketDetail.objects.none()
            )
        else:
            context['detail_formset'] = TicketDetailFormSet(
                queryset=TicketDetail.objects.none()
            )

        # Modal de éxito
        if self.request.GET.get('success') and self.request.GET.get('ticket_id'):
            try:
                ticket = Ticket.objects.get(pk=self.request.GET['ticket_id'])
                context['show_modal'] = True
                context['created_ticket'] = ticket
            except Ticket.DoesNotExist:
                pass

        context['is_edit'] = False

        # Agregar compañía por defecto y su IVA
        from apps.company.models import Company
        company = Company.objects.first()
        if company:
            context['default_company'] = company
            context['iva_percentage'] = company.iva_percentage

        # Breadcrumbs
        context['breadcrumb_list'] = [
            {'label': 'Dashboard', 'url': reverse_lazy('core:dashboard')},
            {'label': 'Tickets', 'url': reverse_lazy('ticket:ticket_list')},
            {'label': 'Nuevo Ticket'}
        ]

        return context

    def form_valid(self, form):
        TicketDetailFormSet = modelformset_factory(
            TicketDetail,
            form=TicketDetailForm,
            extra=1,
            can_delete=True,
        )
        detail_formset = TicketDetailFormSet(
            self.request.POST,
            queryset=TicketDetail.objects.none()
        )
        
        with transaction.atomic():
            # Asignar compañía por defecto
            from apps.company.models import Company
            form.instance.company = Company.objects.first()
            # Copiar IVA de la compañía
            form.instance.iva_percentage = form.instance.company.iva_percentage
            self.object = form.save()

            if detail_formset.is_valid():
                # Guardar detalles
                details = detail_formset.save(commit=False)
                
                # Verificar que hay al menos un detalle
                if not details:
                    messages.error(self.request, 'Debe agregar al menos un producto al ticket.')
                    return self.form_invalid(form)
                
                for detail in details:
                    detail.ticket = self.object
                    detail.save()
                
                # Actualizar total
                self.object.update_total()
                messages.success(self.request, f'Ticket {self.object.document_number} creado exitosamente.')
                
                # Redirigir con parámetros para modal
                return redirect(f"{reverse('ticket:ticket_create')}?success=1&ticket_id={self.object.pk}")
            else:
                messages.error(self.request, 'Error en los detalles del ticket. Verifique los datos.')
                return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error al crear el ticket. Verifique los datos.')
        return super().form_invalid(form)


class TicketUpdateView(UpdateView):
    """
    Vista para editar un ticket existente con sus detalles.
    """
    model = Ticket
    form_class = TicketForm
    template_name = 'ticket/ticket_form.html'
    success_url = reverse_lazy('ticket:ticket_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        TicketDetailFormSet = modelformset_factory(
            TicketDetail,
            form=TicketDetailForm,
            extra=0,  # No extra en edición
            can_delete=True,
        )
        if self.request.POST:
            context['detail_formset'] = TicketDetailFormSet(
                self.request.POST,
                queryset=self.object.details.all()
            )
        else:
            context['detail_formset'] = TicketDetailFormSet(
                queryset=self.object.details.all()
            )
        context['is_edit'] = True
        context['default_company'] = self.object.company
        context['iva_percentage'] = self.object.iva_percentage

        # Breadcrumbs
        context['breadcrumb_list'] = [
            {'label': 'Dashboard', 'url': reverse_lazy('core:dashboard')},
            {'label': 'Tickets', 'url': reverse_lazy('ticket:ticket_list')},
            {'label': f'Editar Ticket #{self.object.document_number}'}
        ]

        return context

    def form_valid(self, form):
        TicketDetailFormSet = modelformset_factory(
            TicketDetail,
            form=TicketDetailForm,
            extra=0,
            can_delete=True,
        )
        detail_formset = TicketDetailFormSet(
            self.request.POST,
            queryset=self.object.details.all()
        )
        
        with transaction.atomic():
            self.object = form.save()

            if detail_formset.is_valid():
                # Eliminar detalles marcados para eliminación
                for deleted_form in detail_formset.deleted_forms:
                    if deleted_form.instance.pk:
                        deleted_form.instance.delete()
                
                # Guardar detalles actualizados y nuevos
                details = detail_formset.save(commit=False)
                
                # Verificar que hay al menos un detalle
                remaining_details = self.object.details.count()
                new_details = len([d for d in details if not d.pk])
                deleted_details = len(detail_formset.deleted_forms)
                
                if remaining_details - deleted_details + new_details <= 0:
                    messages.error(self.request, 'Debe mantener al menos un producto en el ticket.')
                    return self.form_invalid(form)
                
                for detail in details:
                    detail.ticket = self.object
                    detail.save()
                
                # Actualizar total
                self.object.update_total()
                messages.success(self.request, f'Ticket {self.object.document_number} actualizado exitosamente.')
                return redirect(self.success_url)
            else:
                messages.error(self.request, 'Error en los detalles del ticket. Verifique los datos.')
                return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error al actualizar el ticket. Verifique los datos.')
        return super().form_invalid(form)


class TicketDeleteView(DeleteView):
    """
    Vista para eliminar un ticket.
    """
    model = Ticket
    template_name = 'ticket/ticket_confirm_delete.html'
    success_url = reverse_lazy('ticket:ticket_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Ticket eliminado exitosamente.')
        return super().delete(request, *args, **kwargs)


class TicketPrintView(DetailView):
    """
    Vista para imprimir ticket en diferentes formatos.
    """
    model = Ticket
    template_name = 'ticket/ticket_print.html'
    context_object_name = 'ticket'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['details'] = self.object.details.all()
        # Opciones: 58, 80, half (media hoja A4), A4
        context['size'] = self.request.GET.get('size', 'half')
        return context


class TicketMassPrintView(ListView):
    """
    Vista para imprimir múltiples tickets en masa (hasta 6 por página).
    """
    model = Ticket
    template_name = 'ticket/ticket_mass_print.html'
    context_object_name = 'tickets'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('company')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from and date_to:
            queryset = queryset.filter(date__date__range=[date_from, date_to])
        elif date_from:
            queryset = queryset.filter(date__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__date__lte=date_to)
        return queryset.order_by('-date')  # Máximo 6 tickets

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for ticket in context['tickets']:
            ticket.details_list = ticket.details.all()
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


def export_tickets_excel(request):
    """
    Vista para exportar todos los tickets a Excel.
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Encabezados
    headers = [
        'Número de Ticket', 'Fecha', 'Cliente', 'Vendedor', 'CI/RUC', 'Teléfono', 'Placa',
        'Compañía', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal Producto',
        'Subtotal Ticket', 'IVA (%)', 'Monto IVA', 'Total Ticket'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Obtener tickets con detalles
    tickets = Ticket.objects.select_related('company').prefetch_related('details').order_by('-date')
    
    row_num = 2
    for ticket in tickets:
        # Para cada ticket, agregar filas por cada detalle
        for detail in ticket.details.all():
            ws.cell(row=row_num, column=1, value=ticket.document_number)
            ws.cell(row=row_num, column=2, value=ticket.date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=3, value=ticket.client)
            ws.cell(row=row_num, column=4, value=ticket.seller)
            ws.cell(row=row_num, column=5, value=ticket.ci_ruc)
            ws.cell(row=row_num, column=6, value=ticket.phone)
            ws.cell(row=row_num, column=7, value=ticket.plate)
            ws.cell(row=row_num, column=8, value=ticket.company.name if ticket.company else '')
            ws.cell(row=row_num, column=9, value=detail.product)
            ws.cell(row=row_num, column=10, value=float(detail.quantity))
            ws.cell(row=row_num, column=11, value=float(detail.unit_price))
            ws.cell(row=row_num, column=12, value=float(detail.total))
            ws.cell(row=row_num, column=13, value=float(ticket.subtotal))
            ws.cell(row=row_num, column=14, value=float(ticket.iva_percentage))
            ws.cell(row=row_num, column=15, value=float(ticket.iva_amount))
            ws.cell(row=row_num, column=16, value=float(ticket.total))
            row_num += 1

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tickets_export.xlsx'
    
    wb.save(response)
    return response